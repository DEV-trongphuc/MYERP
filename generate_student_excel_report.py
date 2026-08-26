import json
import os
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# Load extracted student data
with open('students_data.json', 'r', encoding='utf-8') as f:
    payload = json.load(f)

students = payload.get('data', [])
total_students = len(students)
print(f"Generating comprehensive Excel report for {total_students} students...")

wb = openpyxl.Workbook()
# Remove default sheet
default_sheet = wb.active

# --- COLOR PALETTE & STYLES ---
FONT_FAMILY = "Segoe UI"
HEADER_FILL = PatternFill(start_color="1E3A8A", end_color="1E3A8A", fill_type="solid") # Dark Navy Blue
HEADER_FONT = Font(name=FONT_FAMILY, size=11, bold=True, color="FFFFFF")
SUBHEADER_FILL = PatternFill(start_color="3B82F6", end_color="3B82F6", fill_type="solid")
SUBHEADER_FONT = Font(name=FONT_FAMILY, size=10, bold=True, color="FFFFFF")

TITLE_FONT = Font(name=FONT_FAMILY, size=16, bold=True, color="1E3A8A")
SECTION_FONT = Font(name=FONT_FAMILY, size=12, bold=True, color="1E3A8A")
BOLD_FONT = Font(name=FONT_FAMILY, size=10, bold=True, color="0F172A")
REGULAR_FONT = Font(name=FONT_FAMILY, size=10, color="1E293B")
MUTED_FONT = Font(name=FONT_FAMILY, size=9, italic=True, color="64748B")

CARD_TITLE_FONT = Font(name=FONT_FAMILY, size=9, bold=True, color="475569")
CARD_VAL_FONT = Font(name=FONT_FAMILY, size=15, bold=True, color="1E3A8A")

ZEBRA_FILL = PatternFill(start_color="F8FAFC", end_color="F8FAFC", fill_type="solid")
WHITE_FILL = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
CARD_BG_BLUE = PatternFill(start_color="EFF6FF", end_color="EFF6FF", fill_type="solid")
CARD_BG_GREEN = PatternFill(start_color="F0FDF4", end_color="F0FDF4", fill_type="solid")
CARD_BG_ORANGE = PatternFill(start_color="FFF7ED", end_color="FFF7ED", fill_type="solid")
CARD_BG_PURPLE = PatternFill(start_color="FAF5FF", end_color="FAF5FF", fill_type="solid")
TOTAL_ROW_FILL = PatternFill(start_color="E2E8F0", end_color="E2E8F0", fill_type="solid")

THIN_BORDER_SIDE = Side(border_style="thin", color="CBD5E1")
THIN_BORDER = Border(left=THIN_BORDER_SIDE, right=THIN_BORDER_SIDE, top=THIN_BORDER_SIDE, bottom=THIN_BORDER_SIDE)
TOTAL_BORDER = Border(top=Side(border_style="thin", color="94A3B8"), bottom=Side(border_style="double", color="1E293B"))

ALIGN_LEFT = Alignment(horizontal="left", vertical="center")
ALIGN_CENTER = Alignment(horizontal="center", vertical="center")
ALIGN_RIGHT = Alignment(horizontal="right", vertical="center")
ALIGN_HEADER = Alignment(horizontal="center", vertical="center", wrap_text=True)

# -------------------------------------------------------------
# 1. SHEET: TỔNG QUAN & DASHBOARD PHÂN TÍCH
# -------------------------------------------------------------
ws_summary = wb.create_sheet(title="📊 TỔNG QUAN PHÂN TÍCH")
ws_summary.views.sheetView[0].showGridLines = True

# Title Header
ws_summary["A1"] = "BÁO CÁO PHÂN TÍCH TOÀN BỘ DỮ LIỆU HỌC VIÊN - IDEAS ERP"
ws_summary["A1"].font = TITLE_FONT
ws_summary["A2"] = f"Thời gian trích xuất: {payload.get('generated_at', '')} | Tổng số hồ sơ phân tích: {total_students:,} học viên"
ws_summary["A2"].font = MUTED_FONT

# Key metrics calculations
total_contract = sum(s.get('contract_value', 0) for s in students)
total_paid = sum(s.get('deposit_paid', 0) for s in students)
total_remaining = sum(s.get('remaining_due', 0) for s in students)
avg_score = sum(s.get('lead_score', 0) for s in students) / max(1, total_students)
total_activities = sum(s.get('total_activities', 0) for s in students)

# Draw 4 KPI Metric Cards
cards = [
    ("TỔNG SỐ HỌC VIÊN", f"{total_students:,}", "A4", "B5", CARD_BG_BLUE, "1E3A8A"),
    ("TỔNG HỌC PHÍ / DOANH THU", f"{total_contract:,.0f} đ", "D4", "E5", CARD_BG_GREEN, "15803D"),
    ("ĐÃ THU / THỰC THU", f"{total_paid:,.0f} đ", "G4", "H5", CARD_BG_ORANGE, "C2410C"),
    ("CÔNG NỢ CÒN PHẢI THU", f"{total_remaining:,.0f} đ", "J4", "K5", CARD_BG_PURPLE, "7E22CE"),
]

for title, val, top_left, bottom_right, bg_fill, color_hex in cards:
    ws_summary.merge_cells(f"{top_left}:{bottom_right}")
    cell = ws_summary[top_left]
    cell.value = f"{title}\n{val}"
    cell.font = Font(name=FONT_FAMILY, size=13, bold=True, color=color_hex)
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell.fill = bg_fill
    
    # Border for the card
    col_start = top_left[0]
    row_start = int(top_left[1:])
    col_end = bottom_right[0]
    row_end = int(bottom_right[1:])
    for r in range(row_start, row_end + 1):
        for c_char in [col_start, col_end]:
            col_idx = openpyxl.utils.column_index_from_string(c_char)
            ws_summary.cell(row=r, column=col_idx).border = THIN_BORDER

# Breakdown by Source
ws_summary["A8"] = "1. PHÂN BỔ HỌC VIÊN THEO NGUỒN TUYỂN SINH (LEAD SOURCE)"
ws_summary["A8"].font = SECTION_FONT

source_stats = {}
for s in students:
    src = s.get('source') or 'Chưa xác định'
    if src not in source_stats:
        source_stats[src] = {'count': 0, 'contract': 0, 'paid': 0, 'debt': 0}
    source_stats[src]['count'] += 1
    source_stats[src]['contract'] += s.get('contract_value', 0)
    source_stats[src]['paid'] += s.get('deposit_paid', 0)
    source_stats[src]['debt'] += s.get('remaining_due', 0)

headers_src = ["Nguồn Tuyển Sinh", "Số Lượng Học Viên", "Tỷ Trọng (%)", "Tổng Doanh Thu Hợp Đồng", "Đã Thu", "Công Nợ Còn Lại"]
for col_i, h_text in enumerate(headers_src, start=1):
    c = ws_summary.cell(row=10, column=col_i, value=h_text)
    c.font = HEADER_FONT
    c.fill = HEADER_FILL
    c.alignment = ALIGN_HEADER
    c.border = THIN_BORDER

row_idx = 11
sorted_sources = sorted(source_stats.items(), key=lambda x: x[1]['count'], reverse=True)
for src, stat in sorted_sources:
    pct = stat['count'] / max(1, total_students)
    fill = ZEBRA_FILL if row_idx % 2 == 0 else WHITE_FILL
    
    ws_summary.cell(row=row_idx, column=1, value=src).alignment = ALIGN_LEFT
    ws_summary.cell(row=row_idx, column=2, value=stat['count']).number_format = '#,##0'
    ws_summary.cell(row=row_idx, column=3, value=pct).number_format = '0.0%'
    ws_summary.cell(row=row_idx, column=4, value=stat['contract']).number_format = '#,##0 "đ"'
    ws_summary.cell(row=row_idx, column=5, value=stat['paid']).number_format = '#,##0 "đ"'
    ws_summary.cell(row=row_idx, column=6, value=stat['debt']).number_format = '#,##0 "đ"'
    
    for c_i in range(1, 7):
        cell = ws_summary.cell(row=row_idx, column=c_i)
        cell.font = REGULAR_FONT
        cell.fill = fill
        cell.border = THIN_BORDER
        if c_i in [2, 3, 4, 5, 6]:
            cell.alignment = ALIGN_RIGHT
    row_idx += 1

# Total Source Row
ws_summary.cell(row=row_idx, column=1, value="TỔNG CỘNG").alignment = ALIGN_LEFT
ws_summary.cell(row=row_idx, column=2, value=total_students).number_format = '#,##0'
ws_summary.cell(row=row_idx, column=3, value=1.0).number_format = '0.0%'
ws_summary.cell(row=row_idx, column=4, value=total_contract).number_format = '#,##0 "đ"'
ws_summary.cell(row=row_idx, column=5, value=total_paid).number_format = '#,##0 "đ"'
ws_summary.cell(row=row_idx, column=6, value=total_remaining).number_format = '#,##0 "đ"'
for c_i in range(1, 7):
    cell = ws_summary.cell(row=row_idx, column=c_i)
    cell.font = BOLD_FONT
    cell.fill = TOTAL_ROW_FILL
    cell.border = TOTAL_BORDER
    if c_i >= 2:
        cell.alignment = ALIGN_RIGHT

# Breakdown by Stage
row_idx += 3
ws_summary.cell(row=row_idx, column=1, value="2. PHÂN BỔ THEO GIAI ĐOẠN ĐÀO TẠO & HỒ SƠ (PIPELINE STAGES)").font = SECTION_FONT
row_idx += 2

stage_stats = {}
for s in students:
    stg = s.get('stage_name') or 'Chưa phân giai đoạn'
    if stg not in stage_stats:
        stage_stats[stg] = {'count': 0, 'contract': 0, 'paid': 0, 'debt': 0}
    stage_stats[stg]['count'] += 1
    stage_stats[stg]['contract'] += s.get('contract_value', 0)
    stage_stats[stg]['paid'] += s.get('deposit_paid', 0)
    stage_stats[stg]['debt'] += s.get('remaining_due', 0)

headers_stg = ["Giai Đoạn Hồ Sơ / Học Tập", "Số Lượng Học Viên", "Tỷ Trọng (%)", "Tổng Doanh Thu Hợp Đồng", "Đã Thu", "Công Nợ Còn Lại"]
for col_i, h_text in enumerate(headers_stg, start=1):
    c = ws_summary.cell(row=row_idx, column=col_i, value=h_text)
    c.font = HEADER_FONT
    c.fill = HEADER_FILL
    c.alignment = ALIGN_HEADER
    c.border = THIN_BORDER

row_idx += 1
sorted_stages = sorted(stage_stats.items(), key=lambda x: x[1]['count'], reverse=True)
for stg, stat in sorted_stages:
    pct = stat['count'] / max(1, total_students)
    fill = ZEBRA_FILL if row_idx % 2 == 0 else WHITE_FILL
    
    ws_summary.cell(row=row_idx, column=1, value=stg).alignment = ALIGN_LEFT
    ws_summary.cell(row=row_idx, column=2, value=stat['count']).number_format = '#,##0'
    ws_summary.cell(row=row_idx, column=3, value=pct).number_format = '0.0%'
    ws_summary.cell(row=row_idx, column=4, value=stat['contract']).number_format = '#,##0 "đ"'
    ws_summary.cell(row=row_idx, column=5, value=stat['paid']).number_format = '#,##0 "đ"'
    ws_summary.cell(row=row_idx, column=6, value=stat['debt']).number_format = '#,##0 "đ"'
    
    for c_i in range(1, 7):
        cell = ws_summary.cell(row=row_idx, column=c_i)
        cell.font = REGULAR_FONT
        cell.fill = fill
        cell.border = THIN_BORDER
        if c_i in [2, 3, 4, 5, 6]:
            cell.alignment = ALIGN_RIGHT
    row_idx += 1

# Auto width for Summary
for col in ws_summary.columns:
    max_len = 0
    col_letter = get_column_letter(col[0].column)
    for cell in col:
        val_str = str(cell.value or '')
        if '\n' in val_str:
            val_str = max(val_str.split('\n'), key=len)
        max_len = max(max_len, len(val_str))
    ws_summary.column_dimensions[col_letter].width = max(max_len + 4, 15)

# -------------------------------------------------------------
# 2. SHEET: DANH SÁCH TOÀN BỘ HỌC VIÊN (DETAILED MASTER DATA)
# -------------------------------------------------------------
ws_data = wb.create_sheet(title="🎓 DANH SÁCH HỌC VIÊN CHI TIẾT")
ws_data.views.sheetView[0].showGridLines = True

# Discover all unique custom field labels across students
all_custom_keys = []
for s in students:
    cf = s.get('custom_fields') or {}
    for k in cf.keys():
        if k not in all_custom_keys:
            all_custom_keys.append(k)

master_headers = [
    ("STT", "number"),
    ("Mã Học Viên (ID)", "number"),
    ("Họ và Tên", "text"),
    ("Giới Tính", "center"),
    ("Ngày Sinh", "center"),
    ("Số Điện Thoại", "text"),
    ("Email", "text"),
    ("CCCD / Passport", "text"),
    ("Địa Chỉ", "text"),
    ("Tỉnh / Thành Phố", "text"),
    ("Quận / Huyện", "text"),
    ("Nghề Nghiệp / Chức Danh", "text"),
    ("Cơ Quan / Doanh Nghiệp", "text"),
    
    # Academic & Program
    ("Chương Trình / Khóa Học", "text"),
    ("Mã Khóa / Dự Án", "center"),
    ("Giai Đoạn Tuyển Sinh / Đào Tạo", "text"),
    ("Trạng Thái Hồ Sơ", "center"),
    ("Phân Loại Khách Hàng", "center"),
    ("Mức Độ Quan Tâm (Nhiệt độ)", "center"),
    ("Điểm Đánh Giá (Lead Score)", "number"),
    ("Xếp Hạng (Grade)", "center"),
    
    # Attribution
    ("Nguồn Tuyển Sinh", "text"),
    ("Chuyên Viên Tư Vấn (Sale)", "text"),
    ("Email Sale", "text"),
    ("Đội Ngũ (Team)", "text"),
    ("Người Cộng Tác", "text"),
    
    # Financials
    ("Mã Phiếu Đặt Cọc", "text"),
    ("Học Phí / Giá Trị Hợp Đồng (VNĐ)", "currency"),
    ("Đã Thanh Toán / Đặt Cọc (VNĐ)", "currency"),
    ("Công Nợ Còn Phải Đóng (VNĐ)", "currency"),
    ("Tiến Độ Đóng Học Phí (%)", "percent"),
    ("Số Phiếu Đặt Cọc", "number"),
    
    # Interactions & Care
    ("Tổng Lượt Chăm Sóc", "number"),
    ("Số Cuộc Gọi", "number"),
    ("Số Cuộc Hẹn", "number"),
    ("Tương Tác Gần Nhất", "center"),
    ("Ghi Chú Học Viên / Chăm Sóc", "text"),
    ("Thẻ Phân Loại (Tags)", "text"),
    ("Ngày Nhập Hệ Thống", "center"),
    ("Cập Nhật Gần Nhất", "center"),
]

# Append custom field headers
for ck in all_custom_keys:
    master_headers.append((f"[Mở rộng] {ck}", "text"))

# Write Master Header Row
for col_idx, (h_name, _) in enumerate(master_headers, start=1):
    c = ws_data.cell(row=1, column=col_idx, value=h_name)
    c.font = HEADER_FONT
    c.fill = HEADER_FILL
    c.alignment = ALIGN_HEADER
    c.border = THIN_BORDER
ws_data.row_dimensions[1].height = 28

# Write Student Rows
for row_i, s in enumerate(students, start=2):
    fill = ZEBRA_FILL if row_i % 2 == 0 else WHITE_FILL
    
    row_values = [
        row_i - 1, # STT
        s.get('id', ''),
        s.get('full_name', ''),
        s.get('gender', ''),
        s.get('dob', ''),
        s.get('phone', ''),
        s.get('email', ''),
        s.get('citizen_id', ''),
        s.get('address', ''),
        s.get('city', ''),
        s.get('district', ''),
        s.get('job_title', ''),
        s.get('company_name', ''),
        
        # Academic & Program
        s.get('project_name', ''),
        s.get('project_code', ''),
        s.get('stage_name', ''),
        s.get('contact_status', ''),
        s.get('customer_type', ''),
        s.get('temperature', ''),
        s.get('lead_score', 0),
        s.get('lead_grade', 'A'),
        
        # Attribution
        s.get('source', ''),
        s.get('owner_name', ''),
        s.get('owner_email', ''),
        s.get('team_name', ''),
        s.get('collaborators', ''),
        
        # Financials
        s.get('deposit_codes', ''),
        s.get('contract_value', 0),
        s.get('deposit_paid', 0),
        s.get('remaining_due', 0),
        (s.get('payment_progress_percent', 0) / 100.0),
        s.get('total_deposits_count', 0),
        
        # Interactions
        s.get('total_activities', 0),
        s.get('call_count', 0),
        s.get('meeting_count', 0),
        s.get('last_contact', ''),
        s.get('notes', ''),
        s.get('tags', ''),
        s.get('created_at', ''),
        s.get('updated_at', ''),
    ]
    
    # Custom fields
    cfs = s.get('custom_fields') or {}
    for ck in all_custom_keys:
        row_values.append(cfs.get(ck, ''))
    
    for col_i, (val, (_, col_type)) in enumerate(zip(row_values, master_headers), start=1):
        cell = ws_data.cell(row=row_i, column=col_i, value=val)
        cell.font = REGULAR_FONT
        cell.fill = fill
        cell.border = THIN_BORDER
        
        if col_type == 'currency':
            cell.number_format = '#,##0 "đ"'
            cell.alignment = ALIGN_RIGHT
        elif col_type == 'percent':
            cell.number_format = '0.0%'
            cell.alignment = ALIGN_RIGHT
        elif col_type == 'number':
            cell.number_format = '#,##0'
            cell.alignment = ALIGN_RIGHT
        elif col_type == 'center':
            cell.alignment = ALIGN_CENTER
        else:
            cell.alignment = ALIGN_LEFT

# Auto-fit columns for Master Data
for col in ws_data.columns:
    col_letter = get_column_letter(col[0].column)
    max_len = len(str(col[0].value or ''))
    # Sample up to first 100 rows for performance
    for cell in col[:100]:
        val_str = str(cell.value or '')
        if len(val_str) > max_len:
            max_len = len(val_str)
    ws_data.column_dimensions[col_letter].width = min(max(max_len + 3, 11), 45)

# Freeze top row and filter
ws_data.freeze_panes = "D2"
ws_data.auto_filter.ref = ws_data.dimensions

# -------------------------------------------------------------
# 3. SHEET: PHÂN TÍCH THEO CHƯƠNG TRÌNH / DỰ ÁN
# -------------------------------------------------------------
ws_program = wb.create_sheet(title="📚 PHÂN TÍCH CHƯƠNG TRÌNH HỌC")
ws_program.views.sheetView[0].showGridLines = True

ws_program["A1"] = "BẢNG PHÂN TÍCH HỌC VIÊN THEO CHƯƠNG TRÌNH / KHÓA HỌC"
ws_program["A1"].font = TITLE_FONT

prog_stats = {}
for s in students:
    prog = s.get('project_name') or 'Chưa gán khóa học'
    if prog not in prog_stats:
        prog_stats[prog] = {'count': 0, 'contract': 0, 'paid': 0, 'debt': 0, 'scores': []}
    prog_stats[prog]['count'] += 1
    prog_stats[prog]['contract'] += s.get('contract_value', 0)
    prog_stats[prog]['paid'] += s.get('deposit_paid', 0)
    prog_stats[prog]['debt'] += s.get('remaining_due', 0)
    prog_stats[prog]['scores'].append(s.get('lead_score', 100))

headers_prog = [
    "Chương Trình / Khóa Học", 
    "Số Lượng Học Viên", 
    "Tỷ Trọng HV (%)", 
    "Tổng Doanh Thu Học Phí", 
    "Thực Thu / Đã Nộp", 
    "Công Nợ Còn Phải Thu", 
    "Tỷ Lệ Thu Hồi (%)",
    "Điểm Đánh Giá TB"
]

for col_i, h_text in enumerate(headers_prog, start=1):
    c = ws_program.cell(row=3, column=col_i, value=h_text)
    c.font = HEADER_FONT
    c.fill = HEADER_FILL
    c.alignment = ALIGN_HEADER
    c.border = THIN_BORDER
ws_program.row_dimensions[3].height = 25

r_prog = 4
sorted_prog = sorted(prog_stats.items(), key=lambda x: x[1]['count'], reverse=True)
for prog_name, stat in sorted_prog:
    pct_hv = stat['count'] / max(1, total_students)
    recovery_rate = (stat['paid'] / stat['contract']) if stat['contract'] > 0 else 0
    avg_sc = sum(stat['scores']) / max(1, len(stat['scores']))
    fill = ZEBRA_FILL if r_prog % 2 == 0 else WHITE_FILL
    
    ws_program.cell(row=r_prog, column=1, value=prog_name).alignment = ALIGN_LEFT
    ws_program.cell(row=r_prog, column=2, value=stat['count']).number_format = '#,##0'
    ws_program.cell(row=r_prog, column=3, value=pct_hv).number_format = '0.0%'
    ws_program.cell(row=r_prog, column=4, value=stat['contract']).number_format = '#,##0 "đ"'
    ws_program.cell(row=r_prog, column=5, value=stat['paid']).number_format = '#,##0 "đ"'
    ws_program.cell(row=r_prog, column=6, value=stat['debt']).number_format = '#,##0 "đ"'
    ws_program.cell(row=r_prog, column=7, value=recovery_rate).number_format = '0.0%'
    ws_program.cell(row=r_prog, column=8, value=avg_sc).number_format = '0.0'
    
    for c_i in range(1, 9):
        cell = ws_program.cell(row=r_prog, column=c_i)
        cell.font = REGULAR_FONT
        cell.fill = fill
        cell.border = THIN_BORDER
        if c_i >= 2:
            cell.alignment = ALIGN_RIGHT
    r_prog += 1

# Total Row
ws_program.cell(row=r_prog, column=1, value="TỔNG CỘNG").alignment = ALIGN_LEFT
ws_program.cell(row=r_prog, column=2, value=total_students).number_format = '#,##0'
ws_program.cell(row=r_prog, column=3, value=1.0).number_format = '0.0%'
ws_program.cell(row=r_prog, column=4, value=total_contract).number_format = '#,##0 "đ"'
ws_program.cell(row=r_prog, column=5, value=total_paid).number_format = '#,##0 "đ"'
ws_program.cell(row=r_prog, column=6, value=total_remaining).number_format = '#,##0 "đ"'
ws_program.cell(row=r_prog, column=7, value=(total_paid / total_contract) if total_contract > 0 else 0).number_format = '0.0%'
ws_program.cell(row=r_prog, column=8, value=avg_score).number_format = '0.0'
for c_i in range(1, 9):
    cell = ws_program.cell(row=r_prog, column=c_i)
    cell.font = BOLD_FONT
    cell.fill = TOTAL_ROW_FILL
    cell.border = TOTAL_BORDER
    if c_i >= 2:
        cell.alignment = ALIGN_RIGHT

for col in ws_program.columns:
    col_letter = get_column_letter(col[0].column)
    max_len = max(len(str(cell.value or '')) for cell in col)
    ws_program.column_dimensions[col_letter].width = max(max_len + 3, 14)

# -------------------------------------------------------------
# 4. SHEET: PHÂN TÍCH THEO CHUYÊN VIÊN TƯ VẤN & TEAM
# -------------------------------------------------------------
ws_sales = wb.create_sheet(title="👥 HIỆU QUẢ SALE & TEAM")
ws_sales.views.sheetView[0].showGridLines = True

ws_sales["A1"] = "BẢNG PHÂN TÍCH HỌC VIÊN THEO CHUYÊN VIÊN TƯ VẤN & ĐỘI NGŨ"
ws_sales["A1"].font = TITLE_FONT

sale_stats = {}
for s in students:
    owner = s.get('owner_name') or 'Chưa phân bổ'
    team = s.get('team_name') or 'Chưa có nhóm'
    key = (owner, team)
    if key not in sale_stats:
        sale_stats[key] = {'count': 0, 'contract': 0, 'paid': 0, 'debt': 0, 'activities': 0}
    sale_stats[key]['count'] += 1
    sale_stats[key]['contract'] += s.get('contract_value', 0)
    sale_stats[key]['paid'] += s.get('deposit_paid', 0)
    sale_stats[key]['debt'] += s.get('remaining_due', 0)
    sale_stats[key]['activities'] += s.get('total_activities', 0)

headers_sale = [
    "Chuyên Viên Tư Vấn (Sale)", 
    "Đội Ngũ (Team)", 
    "Số Lượng Học Viên", 
    "Tỷ Trọng HV (%)", 
    "Tổng Doanh Số Học Phí", 
    "Doanh Thu Đã Thu", 
    "Công Nợ Quản Lý", 
    "Tỷ Lệ Thu Hồi (%)",
    "Tổng Lượt Chăm Sóc"
]

for col_i, h_text in enumerate(headers_sale, start=1):
    c = ws_sales.cell(row=3, column=col_i, value=h_text)
    c.font = HEADER_FONT
    c.fill = HEADER_FILL
    c.alignment = ALIGN_HEADER
    c.border = THIN_BORDER
ws_sales.row_dimensions[3].height = 25

r_sale = 4
sorted_sales = sorted(sale_stats.items(), key=lambda x: x[1]['count'], reverse=True)
for (owner, team), stat in sorted_sales:
    pct_hv = stat['count'] / max(1, total_students)
    recovery_rate = (stat['paid'] / stat['contract']) if stat['contract'] > 0 else 0
    fill = ZEBRA_FILL if r_sale % 2 == 0 else WHITE_FILL
    
    ws_sales.cell(row=r_sale, column=1, value=owner).alignment = ALIGN_LEFT
    ws_sales.cell(row=r_sale, column=2, value=team).alignment = ALIGN_LEFT
    ws_sales.cell(row=r_sale, column=3, value=stat['count']).number_format = '#,##0'
    ws_sales.cell(row=r_sale, column=4, value=pct_hv).number_format = '0.0%'
    ws_sales.cell(row=r_sale, column=5, value=stat['contract']).number_format = '#,##0 "đ"'
    ws_sales.cell(row=r_sale, column=6, value=stat['paid']).number_format = '#,##0 "đ"'
    ws_sales.cell(row=r_sale, column=7, value=stat['debt']).number_format = '#,##0 "đ"'
    ws_sales.cell(row=r_sale, column=8, value=recovery_rate).number_format = '0.0%'
    ws_sales.cell(row=r_sale, column=9, value=stat['activities']).number_format = '#,##0'
    
    for c_i in range(1, 10):
        cell = ws_sales.cell(row=r_sale, column=c_i)
        cell.font = REGULAR_FONT
        cell.fill = fill
        cell.border = THIN_BORDER
        if c_i >= 3:
            cell.alignment = ALIGN_RIGHT
    r_sale += 1

# Total Row
ws_sales.cell(row=r_sale, column=1, value="TỔNG CỘNG").alignment = ALIGN_LEFT
ws_sales.cell(row=r_sale, column=2, value="").alignment = ALIGN_LEFT
ws_sales.cell(row=r_sale, column=3, value=total_students).number_format = '#,##0'
ws_sales.cell(row=r_sale, column=4, value=1.0).number_format = '0.0%'
ws_sales.cell(row=r_sale, column=5, value=total_contract).number_format = '#,##0 "đ"'
ws_sales.cell(row=r_sale, column=6, value=total_paid).number_format = '#,##0 "đ"'
ws_sales.cell(row=r_sale, column=7, value=total_remaining).number_format = '#,##0 "đ"'
ws_sales.cell(row=r_sale, column=8, value=(total_paid / total_contract) if total_contract > 0 else 0).number_format = '0.0%'
ws_sales.cell(row=r_sale, column=9, value=total_activities).number_format = '#,##0'
for c_i in range(1, 10):
    cell = ws_sales.cell(row=r_sale, column=c_i)
    cell.font = BOLD_FONT
    cell.fill = TOTAL_ROW_FILL
    cell.border = TOTAL_BORDER
    if c_i >= 3:
        cell.alignment = ALIGN_RIGHT

for col in ws_sales.columns:
    col_letter = get_column_letter(col[0].column)
    max_len = max(len(str(cell.value or '')) for cell in col)
    ws_sales.column_dimensions[col_letter].width = max(max_len + 3, 14)

# -------------------------------------------------------------
# 5. SHEET: PHÂN BỔ ĐỊA LÝ & TỈNH THÀNH
# -------------------------------------------------------------
ws_geo = wb.create_sheet(title="📍 PHÂN BỔ ĐỊA LÝ & TỈNH THÀNH")
ws_geo.views.sheetView[0].showGridLines = True

ws_geo["A1"] = "BẢNG PHÂN BỔ HỌC VIÊN THEO TỈNH / THÀNH PHỐ"
ws_geo["A1"].font = TITLE_FONT

city_stats = {}
for s in students:
    city = (s.get('city') or '').strip()
    if not city:
        city = 'Chưa cập nhật tỉnh/thành'
    if city not in city_stats:
        city_stats[city] = {'count': 0, 'contract': 0, 'paid': 0}
    city_stats[city]['count'] += 1
    city_stats[city]['contract'] += s.get('contract_value', 0)
    city_stats[city]['paid'] += s.get('deposit_paid', 0)

headers_geo = ["Tỉnh / Thành Phố", "Số Lượng Học Viên", "Tỷ Trọng (%)", "Tổng Doanh Thu Hợp Đồng", "Đã Thu"]
for col_i, h_text in enumerate(headers_geo, start=1):
    c = ws_geo.cell(row=3, column=col_i, value=h_text)
    c.font = HEADER_FONT
    c.fill = HEADER_FILL
    c.alignment = ALIGN_HEADER
    c.border = THIN_BORDER
ws_geo.row_dimensions[3].height = 25

r_geo = 4
sorted_cities = sorted(city_stats.items(), key=lambda x: x[1]['count'], reverse=True)
for city_name, stat in sorted_cities:
    pct = stat['count'] / max(1, total_students)
    fill = ZEBRA_FILL if r_geo % 2 == 0 else WHITE_FILL
    
    ws_geo.cell(row=r_geo, column=1, value=city_name).alignment = ALIGN_LEFT
    ws_geo.cell(row=r_geo, column=2, value=stat['count']).number_format = '#,##0'
    ws_geo.cell(row=r_geo, column=3, value=pct).number_format = '0.0%'
    ws_geo.cell(row=r_geo, column=4, value=stat['contract']).number_format = '#,##0 "đ"'
    ws_geo.cell(row=r_geo, column=5, value=stat['paid']).number_format = '#,##0 "đ"'
    
    for c_i in range(1, 6):
        cell = ws_geo.cell(row=r_geo, column=c_i)
        cell.font = REGULAR_FONT
        cell.fill = fill
        cell.border = THIN_BORDER
        if c_i >= 2:
            cell.alignment = ALIGN_RIGHT
    r_geo += 1

# Total Row
ws_geo.cell(row=r_geo, column=1, value="TỔNG CỘNG").alignment = ALIGN_LEFT
ws_geo.cell(row=r_geo, column=2, value=total_students).number_format = '#,##0'
ws_geo.cell(row=r_geo, column=3, value=1.0).number_format = '0.0%'
ws_geo.cell(row=r_geo, column=4, value=total_contract).number_format = '#,##0 "đ"'
ws_geo.cell(row=r_geo, column=5, value=total_paid).number_format = '#,##0 "đ"'
for c_i in range(1, 6):
    cell = ws_geo.cell(row=r_geo, column=c_i)
    cell.font = BOLD_FONT
    cell.fill = TOTAL_ROW_FILL
    cell.border = TOTAL_BORDER
    if c_i >= 2:
        cell.alignment = ALIGN_RIGHT

for col in ws_geo.columns:
    col_letter = get_column_letter(col[0].column)
    max_len = max(len(str(cell.value or '')) for cell in col)
    ws_geo.column_dimensions[col_letter].width = max(max_len + 3, 14)

# Remove default empty sheet
if "Sheet" in wb.sheetnames:
    wb.remove(wb["Sheet"])

# Save output excel
output_filename = "DANH_SACH_VA_PHAN_TICH_TOAN_BO_HOC_VIEN_IDEAS.xlsx"
output_path = os.path.join(os.getcwd(), output_filename)
wb.save(output_path)
print(f"Report generated successfully: {output_path}")
print(f"File size: {os.path.getsize(output_path) / 1024:.1f} KB")
